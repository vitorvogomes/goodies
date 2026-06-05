#!/usr/bin/env python3
"""Importa a aba "Movimentação" de relatórios B3 (XLSX) -> asset_operations.

Fonte real: relatórios exportados do Portal do Investidor (B3) + Tesouro Direto.
É a base da futura rotina automatizada (mesmo formato "Movimentação").

Uso:
    uv run python ../scripts/import_b3.py <arquivo_ou_glob...> [--dry-run] [--commit]
    # default é DRY-RUN (não grava). Use --commit para gravar no banco.

Ex.:
    uv run python ../scripts/import_b3.py "../files/b3/extrato mensal/"*.xlsx --dry-run
"""
from __future__ import annotations

import glob
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "api"))

import openpyxl

from config import settings
from db.connection import close_pool, get_pool, init_pool
from engines.portfolio import service
from engines.portfolio.b3_import import (
    parse_b3_movimentacao,
    parse_b3_position_prices,
)
from engines.portfolio.migration import import_operations

_SHEET = "Movimentação"


def _read_movimentacao(path: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if _SHEET not in wb.sheetnames:
        return []
    ws = wb[_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    return rows[1:]  # descarta cabeçalho


def _read_snapshot_prices(path: str) -> dict[str, float]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {
        name: list(wb[name].iter_rows(values_only=True))
        for name in wb.sheetnames
    }
    return parse_b3_position_prices(sheets)


def _opt(flag: str) -> str | None:
    if flag in sys.argv:
        i = sys.argv.index(flag)
        if i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return None


def main() -> None:
    snapshot = _opt("--snapshot")
    args = [
        a
        for i, a in enumerate(sys.argv[1:], start=1)
        if not a.startswith("--") and sys.argv[i - 1] != "--snapshot"
    ]
    commit = "--commit" in sys.argv
    if not args:
        print("Uso: import_b3.py <xlsx...> [--snapshot <xlsx>] [--commit]")
        sys.exit(2)

    paths: list[str] = []
    for a in args:
        paths.extend(sorted(glob.glob(a)))

    print("=" * 70)
    print("IMPORT B3 — aba Movimentação")
    print("=" * 70)

    all_ops = []
    for path in paths:
        rows = _read_movimentacao(path)
        ops = parse_b3_movimentacao(rows)
        all_ops.extend(ops)
        print(f"\n{Path(path).name}: {len(rows)} linhas -> {len(ops)} operações")

    if not all_ops:
        print("\nNenhuma operação encontrada.")
        return

    by_tipo = Counter(o["tipo"] for o in all_ops)
    by_cat = Counter(o["asset_category"] for o in all_ops)
    datas = sorted(o["data_operacao"] for o in all_ops)
    print(f"\nTotal: {len(all_ops)} operações")
    print(f"  período:   {datas[0]} .. {datas[-1]}")
    print(f"  por tipo:  {dict(by_tipo)}")
    print(f"  categoria: {dict(by_cat)}")

    if not commit:
        print("\n(DRY-RUN — nada gravado. Use --commit para importar.)")
        return

    import asyncio

    async def _run() -> None:
        await init_pool(settings.database_url)
        pool = get_pool()
        async with pool.acquire() as conn:
            admin = await conn.fetchrow(
                "SELECT id, email FROM users ORDER BY created_at LIMIT 1"
            )
            if not admin:
                print("[erro] Nenhum usuário no banco.")
                sys.exit(1)
            user_id = str(admin["id"])
            report = await import_operations(
                conn, user_id, all_ops, broker_default="Toro/B3"
            )
            n_prices = 0
            if snapshot:
                prices = _read_snapshot_prices(snapshot)
                for ticker, price in prices.items():
                    await service.upsert_price(conn, ticker, price, source="b3")
                n_prices = len(prices)
            await service.invalidate_xirr_cache(user_id)
            xirr = await service.calculate_portfolio_xirr(conn, user_id)
        await close_pool()
        print(f"\nUsuário: {admin['email']}")
        print(f"Importadas: {report['imported']} | Já existentes: {report['skipped']}")
        print(f"Preços semeados (snapshot): {n_prices}")
        c = xirr["consolidated"]
        print(
            "XIRR consolidado: "
            + (f"{c * 100:.2f}% a.a." if c is not None else "indefinido")
        )

    asyncio.run(_run())
    print("=" * 70)


if __name__ == "__main__":
    main()
